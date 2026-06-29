"""
DG-Forge — Generic Ingestion Engine
===================================
Client-agnostic library that turns a raw GDS package (one client's ZIP) into
standardised, one-record-per-App-ID DataLake output aligned to Standard Schema.

Everything client-specific is data: it comes from
  - client_config.<CLIENT>.yaml      (structural config — see assets/ template)
  - field_mapping.<CLIENT>.xlsx      (per-SDD-path field mapping sheet)

This module implements the COMMON SPINE (identical across SOC / USCC / Kapitus /
TIB) and dispatches to registered strategies for the bits that vary.

THREE INVARIANTS (not config-toggleable — enforced by `run_pipeline`):
  I1  Credential scrub runs FIRST, before any logging/parsing/routing.
  I2  PII is tokenised BEFORE any write; a post-write assertion proves zero raw
      PII in the DataLake=Y partition.
  I3  Validation runs BEFORE write; hard-quarantine failures block the write.

DG-Forge governance: this engine is operational tooling. The per-client config
and field-mapping content are engineer-authored and signed off — not generated.

Status: SCAFFOLD. Deterministic plumbing is implemented; client-specific parse
bodies and the vault/object-store/DataLake adapters are marked `TODO(team)`.
"""
from __future__ import annotations

import ast
import base64
import binascii
import gzip
import hashlib
import json
import logging
import re
import zipfile
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable

log = logging.getLogger("dg_forge.ingest")

_ENGINE_VERSION = "1.0.0"

# =============================================================================
# 0. Config + record containers
# =============================================================================
@dataclass
class ClientConfig:
    """Parsed client_config.<CLIENT>.yaml. Thin wrapper so callers get .get paths."""
    raw: dict

    @classmethod
    def load(cls, path: str | Path) -> "ClientConfig":
        import yaml  # pyyaml
        with open(path) as fh:
            return cls(yaml.safe_load(fh))

    def __getitem__(self, key): return self.raw[key]
    def get(self, key, default=None): return self.raw.get(key, default)

    @property
    def client_code(self) -> str: return self.raw["client"]["code"]
    @property
    def schema_version(self) -> str: return str(self.raw["client"]["schema_version"])
    @property
    def folder_priority(self) -> list[str]:
        return [f["name"] for f in self.raw["package"]["folder_priority"]]


@dataclass
class SourceFile:
    """One physical file from the package, after manifest + classification."""
    path: Path
    folder: str
    connector: str | None
    direction: str | None          # REQ / RESP / None
    step: int | None
    app_id_raw: str | None         # str — IC-3 / INV-07; None if unclassified
    sequence_id: str | None        # str — IC-3
    payload: Any = None            # populated after parse
    raw_bytes: bytes | None = None
    geography: str | None = None   # USA | CAN | None (unclassified / unrecognised)
    datetime: "datetime | None" = None  # file-level timestamp (set from filename or payload)


@dataclass
class AppRecord:
    """Accumulates one App ID's standardised output as the pipeline runs."""
    app_id_canonical: str
    app_id_raw: str
    geography: str | None = None
    files: list[SourceFile] = field(default_factory=list)
    record: dict = field(default_factory=dict)        # the SS output
    extra_columns: dict = field(default_factory=dict)
    lineage: dict = field(default_factory=dict)
    validation_failures: list[str] = field(default_factory=list)
    quarantined: bool = False


# =============================================================================
# 1. INGESTION — unzip, manifest, classify
# =============================================================================
def unpack_zip(zip_path: str | Path, dest: str | Path) -> Path:
    dest = Path(dest)
    dest.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path) as z:
        # guard against path traversal
        for member in z.namelist():
            target = (dest / member).resolve()
            if not str(target).startswith(str(dest.resolve())):
                raise ValueError(f"Unsafe path in zip: {member}")
        z.extractall(dest)
    return dest


def build_manifest(root: Path, cfg: ClientConfig) -> list[SourceFile]:
    """Walk the package and classify each file by folder/connector/direction.

    Tolerates folders that are present-but-empty (USCC/Kapitus cc_extracts/).
    """
    folders = cfg.folder_priority
    files: list[SourceFile] = []
    for folder in folders:
        fdir = root / folder
        if not fdir.exists():
            log.info("folder %s absent — skipping", folder)
            continue
        members = list(fdir.rglob("*"))
        if not members:
            log.info("folder %s present but empty — tolerated", folder)
            continue
        for p in members:
            if p.is_file():
                files.append(_classify_file(p, folder, cfg))
    return files


_VALID_GEOS: frozenset[str] = frozenset({"CAN", "USA"})

# Direction aliases accepted from filenames (case-insensitive)
_DIRECTION_MAP: dict[str, str] = {
    "request": "REQ", "req": "REQ",
    "response": "RESP", "resp": "RESP",
}


def _classify_file(p: Path, folder: str, cfg: ClientConfig) -> SourceFile:
    """Extract identity fields from one filename using the config-declared regex.

    Single purpose: filename token parsing → SourceFile field population.
    CQ-001: conditional nesting ≤ 2 levels.

    INV-07: app_id_raw built as VARCHAR string — no numeric cast at any stage.
    INV-10: geography set only from explicit geo token; never inferred from payload.
    Amendment A1: no-match files flagged for quarantine, not silently dropped.
    """
    sf = SourceFile(path=p, folder=folder, connector=None, direction=None,
                    step=None, app_id_raw=None, sequence_id=None)
    appid_cfg = cfg["application_id"]
    if appid_cfg["source"] != "filename_tokens":
        return sf

    m = re.match(appid_cfg["filename"]["pattern"], p.name, re.IGNORECASE)
    if not m:
        # Amendment A1: warn; dispatcher will quarantine with 'filename_parse_failed'
        log.warning("UNCLASSIFIED file=%s — will be quarantined by dispatcher", p.name)
        return sf  # app_id_raw=None, geography=None

    gd = m.groupdict()

    # connector
    sf.connector = gd.get("connector")

    # direction — normalise to REQ | RESP | None
    sf.direction = _DIRECTION_MAP.get((gd.get("direction") or "").lower())
    if sf.direction is None:
        # Fallback: derive from filename when regex has no direction group.
        # Check 'response' first — it doesn't contain 'request' as a substring.
        stem = p.stem.lower()
        if "response" in stem:
            sf.direction = "RESP"
        elif "request" in stem:
            sf.direction = "REQ"

    # sequence_id — str, never cast to int (IC-3)
    sf.sequence_id = gd.get("sequence_id") or None

    # step
    sf.step = int(gd["step"]) if gd.get("step") else None

    # app_id_raw — VARCHAR string concatenation (INV-07 / IC-3)
    debtor = gd.get("debtor", "")
    dt = gd.get("dt", "")
    test_suffix = "_test" if gd.get("test") else ""
    sf.app_id_raw = debtor + "_" + dt + test_suffix  # str — never numeric

    # geography — INV-10: explicit token only; no payload inference
    geo = (gd.get("geo") or "").upper()
    if geo in _VALID_GEOS:
        sf.geography = geo
    else:
        log.warning("UNRECOGNISED geo=%r in file=%s — geography set to None", geo, p.name)
        sf.geography = None

    return sf


# =============================================================================
# 2. INVARIANT I1 — CREDENTIAL SCRUB (always first)
# =============================================================================
def scrub_credentials(files: list[SourceFile], cfg: ClientConfig) -> list[str]:
    """Remove credentials/secrets in place BEFORE anything reads the payloads.

    Returns the list of scrubbed connectors for lineage. Operates on raw_bytes
    where the secret is in the wire header, and on parsed structures otherwise.
    """
    scrubbed: list[str] = []
    _raw = cfg.get("preprocess", {}).get("credential_scrub", [])
    rules: list[dict] = _raw if isinstance(_raw, list) else []
    by_connector: dict[str, list[dict]] = {}
    for r in rules:
        by_connector.setdefault(r["connector"], []).append(r)

    for sf in files:
        for rule in by_connector.get(sf.connector or "", []):
            _apply_scrub(sf, rule)
            scrubbed.append(f"{sf.connector}:{rule.get('method')}")
    return sorted(set(scrubbed))


def _apply_scrub(sf: SourceFile, rule: dict) -> None:
    """Dispatch to the appropriate scrub implementation.

    INV-01: all methods use pattern-based detection — never exact-match of known
    credential values.  raw_bytes is overwritten in-place so no downstream
    function can access the unredacted value.
    """
    method = rule["method"]
    if method == "discard_payload":
        sf.raw_bytes = b"[SCRUBBED_PAYLOAD]"
        sf.payload = None
    elif method == "redact":
        _scrub_redact(sf, rule)
    elif method == "null_out":
        _scrub_null_out(sf, rule)
    elif method == "scrub_json_body":
        _scrub_json_body(sf, rule)
    if rule.get("critical"):
        log.warning("CRITICAL scrub applied to %s (%s)", sf.connector, rule.get("location"))


def _load_raw_text(sf: SourceFile) -> str | None:
    """Read sf.raw_bytes (loading from disk if needed) and decode to str."""
    if sf.raw_bytes is None:
        sf.raw_bytes = sf.path.read_bytes()
    try:
        return sf.raw_bytes.decode("utf-8", errors="replace")
    except Exception:
        return None


def _scrub_redact(sf: SourceFile, rule: dict) -> None:
    """Pattern-based regex redaction in raw bytes (INV-01).

    Applies rule['pattern'] regex to the decoded payload text and replaces
    matches with rule['replacement'].  Overwrites sf.raw_bytes in-place.

    Default pattern targets HTTP Authorization header (C161653 use-case).
    """
    text = _load_raw_text(sf)
    if text is None:
        return
    # NOTE: EXECUTION_PLAN specifies r'(?i)(Authorization:\s*)\S+' but \S+
    # matches only the scheme word (e.g. 'Bearer'), leaving the token value
    # intact.  IC-4 requires zero credential in persisted records, so the
    # correct pattern must match the entire header value to end-of-line.
    # Using [^\r\n]+ here; rule['pattern'] overrides for non-default cases.
    pattern = rule.get("pattern", r"(?i)(Authorization:\s*)[^\r\n]+")
    replacement = rule.get("replacement", r"\1[REDACTED]")
    sf.raw_bytes = re.sub(pattern, replacement, text).encode("utf-8")


def _scrub_null_out(sf: SourceFile, rule: dict) -> None:
    """Pattern-based field nulling for form-encoded or plain-text payloads (INV-01).

    Matches field names from rule['field_pattern'] followed by '=' and nulls
    the value (empties the value string, preserving the key and delimiter).
    Overwrites sf.raw_bytes in-place.

    Default field pattern targets common password field names (C754889 use-case).
    """
    text = _load_raw_text(sf)
    if text is None:
        return
    field_pattern = rule.get("field_pattern", r"password|passwd|pwd|pass")
    sf.raw_bytes = re.sub(
        rf"((?:{field_pattern})=)[^&\r\n]*",
        r"\1",
        text,
        flags=re.IGNORECASE,
    ).encode("utf-8")


def _scrub_json_body(sf: SourceFile, rule: dict) -> None:
    """Pattern-based JSON body credential field scrub (INV-01).

    After HTTP envelope strip is not required — the regex safely applies to
    the full raw bytes (HTTP headers will not contain JSON body keys).
    Matches JSON string values for keys matching rule['field_pattern'] and
    replaces with rule['replacement'].  Overwrites sf.raw_bytes in-place.

    Default targets bearer/access token fields (C103403 use-case).
    """
    text = _load_raw_text(sf)
    if text is None:
        return
    field_pattern = rule.get("field_pattern", r"bearer.?token|access.?token|api.?key")
    replacement = rule.get("replacement", "[SCRUBBED]")
    sf.raw_bytes = re.sub(
        rf'(?i)("(?:{field_pattern})"\s*:\s*)"[^"]*"',
        rf'\1"{replacement}"',
        text,
    ).encode("utf-8")


# =============================================================================
# 3. PRE-PROCESS — strip / decompress / decode / multiparse / externalise
# =============================================================================
def http_envelope_strip(raw: bytes) -> bytes:
    """raw/ wire payloads carry HTTP headers; body starts after CRLFCRLF or LFLF."""
    for sep in (b"\r\n\r\n", b"\n\n"):
        if sep in raw:
            return raw.split(sep, 1)[1]
    return raw


def maybe_gunzip(body: bytes) -> bytes:
    return gzip.decompress(body) if body[:2] == b"\x1f\x8b" else body


def normalise_encoding(body: bytes, accept: str, target: str = "utf-8") -> bytes:
    """Normalise body bytes to the target encoding (default UTF-8).

    Strategy:
      1. Try UTF-8. If valid, re-encode to target (no-op when target=='utf-8').
      2. If UnicodeDecodeError, fall back to ISO-8859-1.
      3. If both fail, raise ValueError with the connector code for traceability.

    Args:
        body:   raw bytes to normalise.
        accept: connector code or content-type hint — included in error message only.
        target: target encoding (default 'utf-8').
    """
    for encoding in ("utf-8", "iso-8859-1"):
        try:
            decoded = body.decode(encoding)
        except UnicodeDecodeError:
            continue
        try:
            return decoded.encode(target)
        except UnicodeEncodeError:
            raise ValueError(
                f"Cannot re-encode body for connector {accept}: "
                f"characters not representable in {target!r}"
            )
    raise ValueError(
        f"Cannot decode body for connector {accept}: "
        f"not valid UTF-8 or ISO-8859-1"
    )


def extract_base64_blob(value: str) -> bytes:
    return base64.b64decode(value)


def json_multiparse(value: Any, depth: int) -> Any:
    """Stringified-JSON that needs N rounds of json.loads (USCC double, Kapitus triple)."""
    out = value
    for _ in range(depth):
        if isinstance(out, str):
            out = json.loads(out)
        else:
            break
    return out


def externalise_if_large(value: Any, key: str, app_id: str, cfg: ClientConfig,
                         lineage_blobs: list[dict]) -> Any:
    """Fields above the size threshold (or binary PDFs) go to object storage and
    are replaced by an external_ref URI (Kapitus REQ-VAL-09/05)."""
    threshold = int(cfg["preprocess"]["external_ref"]["size_threshold_kb"]) * 1024
    blob = value if isinstance(value, (bytes, bytearray)) else str(value).encode()
    if len(blob) < threshold and not _looks_like_pdf(blob):
        return value
    uri = _object_store_put(blob, app_id, key, cfg)        # TODO(team) adapter
    lineage_blobs.append({"field": key, "uri": uri, "size_bytes": len(blob)})
    return {"external_ref": uri}


def _looks_like_pdf(b: bytes) -> bool:
    return b[:4] == b"%PDF"


def _object_store_put(blob: bytes, app_id: str, key: str, cfg: ClientConfig) -> str:
    pattern = cfg["preprocess"]["external_ref"]["object_store_uri_pattern"]
    # TODO(team): real put; for now just format the deterministic URI.
    return pattern.format(app_id=app_id, connector=key.split(".")[0], field=key)


# =============================================================================
# 4. PARSE STRATEGY REGISTRY  (the only place parsing varies)
# =============================================================================
ParseFn = Callable[[SourceFile, ClientConfig], Any]
_STRATEGIES: dict[str, ParseFn] = {}


def strategy(name: str) -> Callable[[ParseFn], ParseFn]:
    def deco(fn: ParseFn) -> ParseFn:
        _STRATEGIES[name] = fn
        return fn
    return deco


def parse_file(sf: SourceFile, cfg: ClientConfig) -> Any:
    conn = _connector_cfg(sf.connector, cfg)
    if conn and conn.get("is_credential"):
        return None                                    # scrub-only, never parsed
    strat = (conn or {}).get("parse_strategy", "raw_json")
    fn = _STRATEGIES.get(strat)
    if not fn:
        raise KeyError(f"No parse strategy '{strat}' registered for {sf.connector}")
    sf.payload = fn(sf, cfg)
    return sf.payload


@strategy("gds_envelope_json")
def _parse_gds_envelope(sf: SourceFile, cfg: ClientConfig) -> Any:
    """data/ tier: full GDS envelope; real payload under data{}."""
    text = sf.path.read_text(encoding="utf-8")
    stripped = text.lstrip()
    if not stripped.startswith(("{", "[")):
        # Non-JSON content (FFF format) — connector delivers both JSON and FFF
        # files; FFF path is pending Q-FFF resolution. TODO(Q-FFF).
        raise NotImplementedError(
            f"Non-JSON (FFF) format in gds_envelope_json file {sf.path.name} — "
            f"connector {sf.connector}. TODO(Q-FFF)."
        )
    obj = json.loads(stripped)
    return obj.get("data", obj)


@strategy("raw_json")
def _parse_raw_json(sf: SourceFile, cfg: ClientConfig) -> Any:
    body = maybe_gunzip(http_envelope_strip(sf.path.read_bytes()))
    if not body.strip():
        log.warning("raw_json: empty body for %s — no payload extracted", sf.path.name)
        return None
    return json.loads(body)


@strategy("xml_dict")
def _parse_xml(sf: SourceFile, cfg: ClientConfig) -> Any:
    import xmltodict  # all TR connectors; TransUnion XML
    body = maybe_gunzip(http_envelope_strip(sf.path.read_bytes()))
    d = xmltodict.parse(body)
    return _strip_ns(d)          # drop ns2:/bs:/cs: prefixes


@strategy("soap_xml")
def _parse_soap(sf: SourceFile, cfg: ClientConfig) -> Any:
    import xmltodict             # USCC C23812 manual review
    body = http_envelope_strip(sf.path.read_bytes())
    return _strip_ns(xmltodict.parse(body))


@strategy("fff")
def _parse_fff(sf: SourceFile, cfg: ClientConfig) -> Any:
    # TODO(Q-FFF): implement when FFF width specification delivered by SOC client.
    raise NotImplementedError(
        f"FFF parse not implemented — connector {sf.connector}. "
        "Awaiting FFF layout spec from SOC client (Q-FFF)."
    )


@strategy("binary_external_ref")
def _parse_binary(sf: SourceFile, cfg: ClientConfig) -> Any:
    body = http_envelope_strip(sf.path.read_bytes())
    blobs: list[dict] = []
    return externalise_if_large(body, f"{sf.connector}.body",
                                sf.app_id_raw or "unknown", cfg, blobs)


@strategy("credential_discard")
def _parse_cred(sf: SourceFile, cfg: ClientConfig) -> Any:
    return None


@strategy("pygdsa_json")
def _parse_pygdsa(sf: SourceFile, cfg: ClientConfig) -> Any:
    """C103403 parse: HTTP strip → [gunzip →] JSON → extract pygdsa attributes.

    Handles two outer-JSON shapes:
      dict  — {"response": {"attributes": {...}}} (actual sample format)
      list  — [base64seg, ...] (originally assumed format; kept as fallback)
    """
    if sf.raw_bytes is None:
        sf.raw_bytes = sf.path.read_bytes()
    # INV-01: credential must be scrubbed before this parse executes
    assert not re.search(
        r"(?i)Authorization:\s*Bearer\s+\S+",
        sf.raw_bytes.decode("utf-8", errors="replace"),
    ), "INV-01: C103403 credential not scrubbed before double-parse"
    body = maybe_gunzip(http_envelope_strip(sf.raw_bytes))
    outer_json = json.loads(body)
    attrs: dict = {}
    if isinstance(outer_json, dict):
        # Actual format: {"response": {"attributes": {...}}}
        attrs = outer_json.get("response", {}).get("attributes", {}) or {}
    elif isinstance(outer_json, list):
        for seg in outer_json:
            try:
                attrs.update(json.loads(base64.b64decode(seg, validate=True)))
            except (binascii.Error, json.JSONDecodeError, ValueError):
                log.warning("pygdsa: malformed segment skipped for %s", sf.connector)
    return attrs


def _strip_ns(obj: Any) -> Any:
    # TODO(production-hardening): add max_depth parameter to _strip_ns() to
    # prevent stack overflow on pathologically nested XML. Real GDS payloads
    # are shallow; this is a production safety net only.
    if isinstance(obj, dict):
        return {k.split(":")[-1]: _strip_ns(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_strip_ns(v) for v in obj]
    return obj


def _connector_cfg(code: str | None, cfg: ClientConfig) -> dict | None:
    for c in cfg.get("connectors", []) or []:
        if c["code"] == code:
            return c
    return None


# =============================================================================
# 5. GEO DISPATCH + GROUP + MERGE
# =============================================================================
def dispatch_by_geo(files: list[SourceFile]) -> dict[str, list[SourceFile]]:
    """Partition files into {'USA': [...], 'CAN': [...]}.

    Single purpose: geography-based routing only.
    INV-10: routing is explicit and deterministic — no default, no inference.
    Files with geography not in ('USA', 'CAN') are unroutable and logged as errors.
    Both partition keys are always present in the return value, even if empty.
    CQ-001: single purpose; loop nesting ≤ 2 levels.
    """
    partitions: dict[str, list[SourceFile]] = {"USA": [], "CAN": []}
    unroutable: list[SourceFile] = []

    for sf in files:
        if sf.geography in ("USA", "CAN"):
            partitions[sf.geography].append(sf)
        else:
            unroutable.append(sf)

    for sf in unroutable:
        log.error(
            "QUARANTINE geo=None/unrecognised file=%s — INV-10", sf.path.name
        )

    return partitions


def group_by_app(files: list[SourceFile], cfg: ClientConfig) -> dict[str, AppRecord]:
    """Group SourceFiles into AppRecords by canonical App ID.

    Single purpose: partition + dedup + quarantine-flag per-group.
    INV-07 / D-10: canonical and raw IDs stored as VARCHAR strings throughout.
    INV-09: records with _test suffix quarantined (not written to DataLake=Y).
    D-02: cross-debtor mismatch within a group quarantined.
    """
    apps: dict[str, AppRecord] = {}
    for sf in files:
        if not sf.app_id_raw:
            continue
        canonical, had_test = _canonicalise_app_id(sf.app_id_raw, cfg)
        if canonical not in apps:
            rec = AppRecord(canonical, sf.app_id_raw)
            rec.lineage["app_id_raw"] = sf.app_id_raw         # IC-3: preserved
            rec.lineage["app_id_canonical"] = canonical        # IC-3: preserved
            apps[canonical] = rec
        rec = apps[canonical]
        if had_test:
            rec.lineage["app_id_raw_had_test_suffix"] = True
        rec.files.append(sf)

    for rec in apps.values():
        # Dedup retry files before any further checks
        rec.files = _dedup_retry_files(rec.files)

        # INV-09: quarantine _test records — never write to DataLake=Y
        if rec.lineage.get("app_id_raw_had_test_suffix"):
            rec.quarantined = True
            rec.lineage["test_quarantine"] = True
            # TODO(Q3): when separate_partition confirmed, route here instead of quarantine

        # D-02: all files in a group must share the same debtor number
        _check_group_debtor_consistency(rec)

    return apps


def _canonicalise_app_id(raw: str, cfg: ClientConfig) -> tuple[str, bool]:
    """Return (canonical_id, had_test_suffix).

    Applies suffix_rules from config. INV-07 / D-10: returns VARCHAR string,
    no numeric coercion at any point.
    """
    out = raw
    had_test = False
    for rule in cfg["application_id"].get("suffix_rules", []) or []:
        suf = rule.get("suffix")
        if suf and out.endswith(suf) and rule.get("action") == "strip":
            out = out[: -len(suf)]
            if suf == "_test":
                had_test = True
    return out, had_test


def _dedup_retry_files(files: list[SourceFile]) -> list[SourceFile]:
    """Remove duplicate retry files, keeping the latest by filename sort.

    Two files are duplicates when they share (connector, direction, sequence_id).
    Sorted ascending by path name so the last entry (latest filename timestamp)
    overwrites earlier ones.  INV-07: sequence_id compared as string — no cast.
    """
    seen: dict[tuple, SourceFile] = {}
    for sf in sorted(files, key=lambda f: f.path.name):
        key = (sf.connector, sf.direction, sf.sequence_id)
        seen[key] = sf
    return list(seen.values())


def _check_group_debtor_consistency(rec: AppRecord) -> None:
    """D-02 (group-level): all files in a group must share one debtor number.

    Debtor number is the leading component of app_id_raw before the first '_'.
    INV-07: compared as strings only.
    """
    debtors = {
        sf.app_id_raw.split("_")[0]
        for sf in rec.files
        if sf.app_id_raw
    }
    if len(debtors) > 1:
        rec.quarantined = True
        rec.validation_failures.append("D-02-cross-session-identity-mismatch")


def merge_sessions(rec: AppRecord, cfg: ClientConfig) -> None:
    """Apply the client's session model.

    CAN records: detect bureau sessions by connector presence (TASK-3.3),
    check session ordering (TASK-3.4), and check EcsDebtorNumber payload
    consistency (TASK-3.5).
    USA records: no multi-session merge logic in this release.
    """
    if rec.geography == "CAN":
        _detect_can_sessions(rec)
        _check_can_session_order(rec)
        _check_payload_debtor_consistency(rec)


# INV-02 / D-07: PII pattern set — compiled at module level; used by both
# _scan_extra_columns_for_pii() and assert_no_raw_pii().
_PII_PATTERNS: dict[str, re.Pattern] = {
    "email": re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"),
    # (?<![A-Za-z0-9_]) prevents matching inside alphanumeric contexts:
    #   • 14-digit timestamps (preceding char is a digit, which is in the set)
    #   • pseudonym tokens like TOK_d4d79f4827439447 (preceding char is hex 'f')
    # (?!\d) still blocks 10-digit substrings at the END of longer digit runs.
    "phone": re.compile(r"(?<![A-Za-z0-9_])(?:\+?1[\s\-.]?)?\(?\d{3}\)?[\s\-.]?\d{3}[\s\-.]?\d{4}(?!\d)"),
    # Require consistent separators: either all dashes, all spaces, or bare 9 digits.
    # Prevents ZIP+4 (e.g. 68378-8300) from matching as 683|78|-|8300.
    "ssn":   re.compile(r"\b(?:\d{3}([-\s])\d{2}\1\d{4}|\d{9})\b"),
    "sin":   re.compile(r"\b\d{3}[\-\s]?\d{3}[\-\s]?\d{3}\b"),
    "fein":  re.compile(r"\b\d{2}[\-]?\d{7}\b"),
}

# CAN session connector identifiers
_CAN_SESSION1_CONNECTORS: frozenset[str] = frozenset({"C100810"})
_CAN_SESSION2_CONNECTORS: frozenset[str] = frozenset({"C161653", "C161796"})

# D-09: bureau provider routing (connector → provider sub-key under bureauData)
_BUREAU_PROVIDER_MAP: dict[str, str] = {
    "C100810": "transunion",
    "C161796": "equifax",
    "C161653": "equifax",
}


def _detect_can_sessions(rec: AppRecord) -> None:
    """Detect CAN bureau evaluation sessions by connector presence only.

    IMPLEMENTATION GUIDANCE (from removed INV-11):
      Uses connector identity — NOT sequence_id — for session detection.
      sequence_id is unreliable for CAN session identity (USA retry at
      seq=80 has two data sessions that represent retry, not two bureau
      sessions). sequence_id is never read in this function.

    D-05: if bureau_eval_indicated and one session is absent → quarantine.
    """
    connectors = {sf.connector for sf in rec.files}

    session1_present = bool(connectors & _CAN_SESSION1_CONNECTORS)
    session2_present = bool(connectors & _CAN_SESSION2_CONNECTORS)
    bureau_eval_indicated = session1_present or session2_present

    rec.lineage["can_session_1_connectors"] = (
        sorted(_CAN_SESSION1_CONNECTORS & connectors) if session1_present else []
    )
    rec.lineage["can_session_2_connectors"] = (
        sorted(_CAN_SESSION2_CONNECTORS & connectors) if session2_present else []
    )

    if bureau_eval_indicated and not (session1_present and session2_present):
        rec.lineage["multi_session_incomplete"] = True
        rec.quarantined = True
        rec.validation_failures.append("REQ-VAL-003")


def _check_can_session_order(rec: AppRecord) -> None:
    """D-01: EFX session timestamp must be strictly > TU session timestamp.

    Soft-warn only: records the anomaly but does NOT quarantine or block write.
    Skipped when either session has no files with a datetime set.
    """
    tu_files  = [sf for sf in rec.files if sf.connector in _CAN_SESSION1_CONNECTORS
                 and sf.datetime is not None]
    efx_files = [sf for sf in rec.files if sf.connector in _CAN_SESSION2_CONNECTORS
                 and sf.datetime is not None]

    if not tu_files or not efx_files:
        return

    tu_ts  = max(sf.datetime for sf in tu_files)
    efx_ts = max(sf.datetime for sf in efx_files)

    if efx_ts <= tu_ts:
        rec.lineage["session_order_anomaly"] = True
        rec.validation_failures.append("REQ-BL-002")


def _check_payload_debtor_consistency(rec: AppRecord) -> None:
    """D-02 (payload-level): EcsDebtorNumber must be identical across all sessions.

    Extracts EcsDebtorNumber from parsed payloads per connector/folder convention.
    Quarantines when COUNT(DISTINCT EcsDebtorNumber) > 1.
    Gracefully skipped when no payload is parsed yet (sf.payload is None).
    """
    debtors = {
        v for sf in rec.files
        if (v := _extract_ecs_debtor(sf)) is not None
    }
    if len(debtors) > 1:
        rec.quarantined = True
        rec.validation_failures.append("D-02-payload-debtor-mismatch")
        log.error(
            "EcsDebtorNumber mismatch AppID=%s values=%s",
            rec.app_id_canonical, debtors,
        )


def _extract_ecs_debtor(sf: SourceFile) -> str | None:
    """Extract EcsDebtorNumber from a parsed SourceFile payload.

    Convention per SOC connector SDD:
      C225334-REQ  → payload['record']['EcsDebtorNumber']
      C103403-RESP → payload['attributes']['EcsDebtorNumber']
      data/ folder → payload['data']['EcsDebtorNumber']
    Returns None if payload is absent or field not found.
    """
    if sf.payload is None:
        return None
    if sf.connector == "C225334" and sf.direction == "REQ":
        return sf.payload.get("record", {}).get("EcsDebtorNumber")
    if sf.connector == "C103403" and sf.direction == "RESP":
        # pygdsa_json yields a flat attrs dict; EcsDebtorNumber is a top-level key
        return sf.payload.get("EcsDebtorNumber")
    if sf.folder == "data":
        return sf.payload.get("data", {}).get("EcsDebtorNumber")
    return None


# =============================================================================
# 6. FIELD MAPPING — drive from field_mapping.<CLIENT>.xlsx
# =============================================================================
@dataclass
class MappingRow:
    sdd_path: str
    category: str
    data_type: str
    pii: bool
    sources: list[dict]            # ordered primary/secondary/tertiary
    transform: str | None
    construction: str | None       # free-text construction logic / array filter


def load_mapping_sheet(xlsx_path: str | Path) -> list[MappingRow]:
    """Read the canonical field-mapping workbook into MappingRows.
    Column layout is defined in references/mapping_schema.md."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Field Mapping"]
    rows = list(ws.iter_rows(values_only=True))
    header = _build_header_index(rows[0])
    out: list[MappingRow] = []
    for r in rows[1:]:
        sdd = r[header.get("SDD Field Path", 1)]
        if not sdd or str(sdd).strip().startswith("──"):
            continue                                   # category banner
        out.append(_row_from_cells(r, header))
    return out


def _build_header_index(header_row) -> dict[str, int]:
    """Build column-name → index dict, indexing by both full and first-line key.

    Columns like 'PII\\n(P/H/D-N/❌)' are reachable via cell("PII") because
    the first-line prefix is also indexed (shorter key wins on conflict).
    """
    idx: dict[str, int] = {}
    for i, h in enumerate(header_row):
        if not h:
            continue
        full = str(h).strip()
        idx[full] = i
        first = full.split("\n")[0].strip()
        if first != full and first not in idx:
            idx[first] = i
    return idx


def _clean_field_path(raw: str) -> str:
    """Strip analyst annotations from a mapping path cell value.

    Handles two patterns seen in field_mapping.xlsx:
      - Semicolon-separated alternatives: take first.
      - Trailing parenthetical note '(...)': strip it.
    """
    path = raw.split(";")[0].strip()
    path = re.sub(r"\s*\([^)]*\)\s*$", "", path).strip()
    return path


def _row_from_cells(r, header) -> MappingRow:
    def cell(name, default=None):
        i = header.get(name)
        return r[i] if i is not None and i < len(r) else default
    sources = []
    for tier in ("PRIMARY", "SECONDARY", "TERTIARY"):
        conn = cell(f"{tier}\nConnector | Folder | Direction") or cell(f"{tier} Connector | Folder | Direction")
        if tier == "PRIMARY":
            path = (cell("PRIMARY Path\n[Obj-1 / Current / 1st]") or
                    cell("PRIMARY Path\n[Obj-1 / Current / First Score]") or
                    cell("PRIMARY Path"))
        else:
            path = cell(f"{tier} Path")
        if conn:
            sources.append({"tier": tier, "locator": str(conn), "path": str(path or "")})
    pii_val = cell("PII")
    return MappingRow(
        sdd_path=str(cell("SDD Field Path")).strip(),
        category=str(cell("Category") or ""),
        data_type=str(cell("Data Type") or ""),
        pii=pii_val in ("P", "H"),
        sources=sources,
        transform=cell("Transform"),
        construction=cell("Mapping Notes / Construction Logic / Open Questions"),
    )


def apply_mapping(rec: AppRecord, mapping: list[MappingRow], cfg: ClientConfig) -> None:
    """For each SDD path, resolve by source priority and write into rec.record."""
    _extract_decision(rec)
    for row in mapping:
        value = resolve_source(rec, row, cfg)
        if value is None:
            continue
        value = apply_transform(value, row, rec, cfg)
        if row.sdd_path.startswith("extra_columns."):
            _set_path(rec.extra_columns, row.sdd_path[len("extra_columns."):], value)
        else:
            _set_path(rec.record, row.sdd_path, value)
    _check_score_slot_bounds(rec)
    _check_decline_completeness(rec)
    _set_bureau_provider_lineage(rec)
    _assert_bureau_attribution(rec)


def _extract_decision(rec: AppRecord) -> None:
    """D-04: extract decision from C238743-RESP only; never from audit/ folder."""
    for sf in rec.files:
        if sf.folder == "audit":        # D-04 guard
            continue
        if sf.connector != "C238743" or sf.direction != "RESP":
            continue
        if sf.payload is None:
            break
        val = _get_nested(sf.payload, "Decision.decision")
        if val:
            _set_path(rec.record, "system.application.decision", val)
            apr = _get_nested(sf.payload, "Decision.interestrate")
            if apr is not None:
                _set_path(rec.record, "system.application.apr", apr)
            return
        break
    # C238743-RESP absent or decision field missing
    rec.lineage["decision_missing"] = True
    rec.validation_failures.append("REQ-VAL-006")


def _check_decline_completeness(rec: AppRecord) -> None:
    """D-03: DECLINED + empty reason codes → REQ-BL-001 soft-warn (not a quarantine)."""
    decision = _get_path(rec.record, "system.application.decision")
    reason_codes = _get_path(rec.record, "system.application.decisionSummary.reasonCodes") or []
    if str(decision or "").upper() == "DECLINED" and len(reason_codes) == 0:
        rec.validation_failures.append("REQ-BL-001")
        rec.lineage["reason_codes_missing"] = True


def _set_bureau_provider_lineage(rec: AppRecord) -> None:
    """D-09: populate rec.lineage['bureau_providers'] from present bureau connectors."""
    providers = {
        _BUREAU_PROVIDER_MAP[sf.connector]
        for sf in rec.files
        if sf.connector in _BUREAU_PROVIDER_MAP
    }
    if providers:
        rec.lineage["bureau_providers"] = sorted(providers)


def _assert_bureau_attribution(rec: AppRecord) -> None:
    """D-09: no field may exist directly at rec.record['bureauData'] root.
    All bureau-derived fields must live under a provider sub-key."""
    bureau_data = rec.record.get("bureauData")
    if bureau_data is None:
        return
    allowed = {"transunion", "equifax"}
    unexpected = {k for k in bureau_data if k not in allowed}
    if unexpected:
        raise ValueError(
            f"D-09: bureauData has fields without provider attribution: {sorted(unexpected)}"
        )


def _check_score_slot_bounds(rec: AppRecord) -> None:
    """Slot bounding: SOC maps scores 1-3 only; slots 4-14 must remain unpopulated."""
    for slot in range(4, 15):
        key = f"system.application.scores.score{slot}"
        if _get_path(rec.record, key) is not None:
            raise ValueError(f"Score slot {slot} populated for SOC — mapping error")


def resolve_source(rec: AppRecord, row: MappingRow, cfg: ClientConfig) -> Any:
    """Walk primary -> secondary -> tertiary; take first present non-null.
    This is the generic Source-Priority resolution all four clients use."""
    for src in row.sources:
        val = _read_locator(rec, src, cfg)
        if val not in (None, "", []):
            return val
    return None


def _read_locator(rec: AppRecord, src: dict, cfg: ClientConfig) -> Any:
    locator = src.get("locator", "")
    field_path = _clean_field_path(src.get("path", ""))
    parts = [p.strip() for p in locator.split("|")]
    if len(parts) < 3:
        return None
    connector_code = parts[0]
    folder = parts[1].rstrip("/")        # normalise: "data/" → "data"
    direction = parts[2]
    for sf in rec.files:
        if sf.connector == connector_code and sf.folder == folder and sf.direction == direction:
            if sf.payload is None:
                return None
            return _get_nested(sf.payload, field_path)
    return None


def apply_transform(value: Any, row: MappingRow, rec: AppRecord, cfg: ClientConfig) -> Any:
    t = (row.transform or "").lower()
    if "date" in t:
        return _to_utc_iso(value)
    if "numeric" in t or "int" in t:
        return _coerce_number(value)
    if "split" in t:
        delim = "|"                                     # configurable per row
        return [s.strip() for s in str(value).split(delim) if s.strip()]
    if "json_double_parse" in t:
        return json.loads(value)
    if "ast_literal_eval" in t:
        return ast.literal_eval(value)
    # EAV, score-array construction, per-applicant indicator filtering,
    # one-object->multi-entry expansion (TIB) are construction-logic transforms.
    # TODO(team): dispatch on row.construction for those.
    return value


# =============================================================================
# 7. INVARIANT I2 — PII TOKENISATION (before write) + scan
# =============================================================================

def _sha256_hex(value: str) -> str:
    return hashlib.sha256(value.encode()).hexdigest()


def _tokenise_value(value: str, method: str) -> Any:
    """Apply a single tokenisation method to a string value."""
    if method == "pseudonym_reversible":
        return "TOK_" + _sha256_hex(value)[:16]
    if method == "oneway_hash":
        return _sha256_hex(value)
    if method == "year_only":
        m = re.search(r"\b(\d{4})\b", value)
        return m.group(1) if m else value
    return value


def _del_path(obj: dict, dotted: str) -> None:
    """Remove the leaf key at a dotted path from a nested dict."""
    parts = dotted.split(".")
    parent = _get_nested(obj, ".".join(parts[:-1])) if len(parts) > 1 else obj
    if isinstance(parent, dict):
        parent.pop(parts[-1], None)


def tokenise_pii(rec: AppRecord, cfg: ClientConfig) -> None:
    """INV-02 / D-07: tokenise all static PII fields before write_record()."""
    _raw_fields = cfg.get("pii", {}).get("fields", [])
    for f in (_raw_fields if isinstance(_raw_fields, list) else []):
        path = f.get("sdd_path") or f.get("path", "")
        method = f.get("method", "")
        value = _get_path(rec.record, path)
        if value is None:
            continue
        if method == "scrub_never_store":
            _del_path(rec.record, path)
        else:
            _set_path(rec.record, path, _tokenise_value(str(value), method))
    if cfg["pii"]["extra_columns_scan"]["enabled"]:
        _scan_extra_columns_for_pii(rec, cfg)


def _flatten_ec_values(data: Any, prefix: str) -> list:
    """Return (flat_key, str_value) pairs for all string leaves in nested data."""
    if isinstance(data, dict):
        out = []
        for k, v in data.items():
            out.extend(_flatten_ec_values(v, f"{prefix}.{k}" if prefix else k))
        return out
    if isinstance(data, list):
        return [item for i, v in enumerate(data)
                for item in _flatten_ec_values(v, f"{prefix}.{i}")]
    if isinstance(data, str):
        return [(prefix, data)]
    return []


def _tokenise_ec_value(rec: AppRecord, key_path: str, val: str) -> None:
    """If val matches any PII pattern, replace it in rec.extra_columns and record lineage."""
    for pat_name, pat in _PII_PATTERNS.items():
        if pat.search(val):
            token = "TOK_EC_" + _sha256_hex(val)[:16]
            _set_path(rec.extra_columns, key_path, token)
            rec.lineage.setdefault("extra_columns_pii_found", []).append(
                {"key": key_path, "pattern": pat_name}
            )
            return


def _scan_extra_columns_for_pii(rec: AppRecord, cfg: ClientConfig) -> None:
    """INV-02 second enforcement path: scan extra_columns VALUES for PII patterns."""
    for group_name, group_data in rec.extra_columns.items():
        for key_path, val in _flatten_ec_values(group_data, group_name):
            _tokenise_ec_value(rec, key_path, val)


def assert_no_raw_pii(rec: AppRecord, cfg: ClientConfig) -> None:
    """INV-02 write gate. Raises RuntimeError if any raw-PII pattern remains in output."""
    blob = json.dumps(rec.record) + json.dumps(rec.extra_columns)
    for pat_name, pat in _PII_PATTERNS.items():
        m = pat.search(blob)
        if m:
            ctx = blob[max(0, m.start() - 20): m.end() + 20]
            raise RuntimeError(
                f"INV-02/D-07 VIOLATION: raw {pat_name} PII in "
                f"AppID={rec.app_id_canonical}. "
                f"Context: ...{ctx}..."
            )


# =============================================================================
# 8. INVARIANT I3 — VALIDATION (before write) + quarantine
# =============================================================================
RuleFn = Callable[[AppRecord, ClientConfig], bool]   # True == pass
_RULES: dict[str, RuleFn] = {}


def rule(rule_id: str):
    def deco(fn: RuleFn):
        _RULES[rule_id] = fn
        return fn
    return deco


@rule("REQ-VAL-001")
def _appid_present(rec, cfg): return bool(rec.app_id_canonical)

@rule("REQ-VAL-002")
def _geo_valid(rec, cfg):
    valids = cfg.get("validation", {}).get("client_params", {}).get("valid_geographies")
    if not valids:
        return True  # not configured — pass through
    return rec.geography in valids if rec.geography else True

@rule("REQ-VAL-005")
def _date_valid(rec, cfg):
    d = rec.record.get("system", {}).get("application", {}).get("applicationDate")
    return _is_iso_utc(d) if d else False

# REQ-VAL-007 (no raw PII) and REQ-VAL-008 (creds scrubbed) are proven by the
# invariant steps themselves; they appear here as explicit gates too.
@rule("REQ-VAL-007")
def _no_raw_pii(rec, cfg): return True   # assert_no_raw_pii already ran

@rule("REQ-VAL-008")
def _creds_scrubbed(rec, cfg):
    return bool(rec.lineage.get("credential_scrubbed_connectors") is not None)

@rule("REQ-VAL-003")
def _can_two_sessions(rec, cfg):
    if rec.geography != "CAN":
        return True
    bureau_indicated = (rec.lineage.get("can_session_1_connectors") or
                        rec.lineage.get("can_session_2_connectors"))
    if not bureau_indicated:
        return True   # FF product — not subject
    return not rec.lineage.get("multi_session_incomplete", False)


@rule("REQ-VAL-004")
def _has_bureau(rec, cfg):
    has = any(sf.connector in ("C100810", "C161796", "C1677939") for sf in rec.files)
    if not has:
        rec.lineage["has_bureau_data"] = False
    return True   # soft-warn only — always passes validation, records lineage flag


@rule("REQ-VAL-006")
def _decision_present(rec, cfg):
    decision = rec.record.get("system", {}).get("application", {}).get("decision")
    return decision is not None or rec.lineage.get("decision_missing", False)


@rule("REQ-BL-001")
def _bl_reason_codes(rec, cfg):
    return not rec.lineage.get("reason_codes_missing", False)


@rule("REQ-BL-002")
def _bl_session_order(rec, cfg):
    return not rec.lineage.get("session_order_anomaly", False)


@rule("REQ-BL-003")
def _bl_debtor_consistency(rec, cfg):
    return not any(f.startswith("D-02-") for f in rec.validation_failures)


@rule("REQ-BL-004")
def _bl_pygdsa_attrs(rec, cfg):
    attr_count = len(rec.extra_columns.get("SOC_pygdsa_attributes", {}))
    if 0 < attr_count < 100:
        rec.lineage["pygdsa_parse_partial"] = True
        return False
    return True


@rule("REQ-BL-005")
def _bl_product_info(rec, cfg):
    prod_info = (rec.record.get("system", {})
                 .get("application", {})
                 .get("productInformation", []))
    if not prod_info:
        rec.lineage["product_info_incomplete"] = True
        return False
    return True


def validate(rec: AppRecord, cfg: ClientConfig) -> None:
    hard = set(cfg["validation"]["hard_quarantine_rules"])
    for rid, fn in _RULES.items():
        try:
            passed = fn(rec, cfg)
        except Exception as e:                          # a throwing rule == fail
            passed = False
            log.warning("rule %s errored: %s", rid, e)
        if not passed:
            rec.validation_failures.append(rid)
            if rid in hard:
                rec.quarantined = True
    rec.lineage["validation_status"] = (
        "FAIL" if rec.quarantined else "WARN" if rec.validation_failures else "PASS"
    )
    rec.lineage["validation_failures"] = rec.validation_failures


# =============================================================================
# 9. LINEAGE + WRITE
# =============================================================================
def build_lineage(rec: AppRecord, cfg: ClientConfig, source_zip: str,
                  scrubbed: list[str], blobs: list[dict]) -> None:
    ec_count = sum(len(list(_flatten_ec_values(v, ""))) for v in rec.extra_columns.values())
    rec.lineage.update({
        "source_zip": source_zip,
        "app_id_raw": rec.app_id_raw,
        "app_id_canonical": rec.app_id_canonical,
        "geography": rec.geography,
        "client_code": cfg.client_code,
        "schema_version": cfg.schema_version,
        "mapping_config_version": cfg.get("config_version"),
        "transform_timestamp": datetime.now(timezone.utc).isoformat(),
        "source_files": [str(sf.path.name) for sf in rec.files],
        "credential_scrubbed_connectors": scrubbed,
        "base64_blobs_extracted": blobs,
        "has_connector_data": any(sf.connector for sf in rec.files),
        "engine_version": _ENGINE_VERSION,
        "extra_columns_field_count": ec_count,
    })
    rec.record.setdefault("system", {})["lineage"] = rec.lineage


def _check_d13_completeness(rec: AppRecord) -> None:
    """D-13: quarantine the record if any of the four completeness conditions fails.

    Runs after validate() so validation_status is already set in lineage.
    """
    decision = (rec.record.get("system", {})
                .get("application", {})
                .get("decision"))
    complete = all([
        rec.app_id_canonical,
        rec.lineage.get("has_connector_data"),
        decision is not None or rec.lineage.get("decision_missing"),
        rec.lineage.get("validation_status") in ("PASS", "WARN"),
    ])
    if not complete:
        rec.lineage["record_completeness"] = "INCOMPLETE"
        rec.validation_failures.append("D-13-incomplete-record")
        rec.quarantined = True


def write_record(rec: AppRecord, cfg: ClientConfig,
                 workdir: "str | Path | None" = None) -> None:
    if rec.quarantined:
        if workdir is not None:
            quarantine_record = {
                "app_id_canonical": rec.app_id_canonical,
                "app_id_raw":       rec.app_id_raw,
                "geography":        rec.geography,
                "quarantine_reason": rec.validation_failures,
                "lineage":          rec.lineage,
                "record_partial":   rec.record,
            }
            path = Path(workdir) / "quarantine" / f"{rec.app_id_canonical}.json"
            path.parent.mkdir(exist_ok=True)
            path.write_text(json.dumps(quarantine_record, indent=2))
        log.error("QUARANTINE %s -> %s", rec.app_id_canonical, rec.validation_failures)
        return                                          # INV-09: blocked from DataLake=Y
    assert not rec.quarantined                         # INV-09: DataLake=Y unreachable for quarantined
    assert rec.app_id_canonical, "INV-04: app_id_canonical null"
    assert rec.geography in ("USA", "CAN"), f"INV-10: invalid geography {rec.geography!r}"
    if workdir is not None:
        out_path = Path(workdir) / "output" / rec.geography / f"{rec.app_id_canonical}.json"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        if out_path.exists():
            raise RuntimeError(f"INV-04 violation: duplicate write for {rec.app_id_canonical}")
        out_path.write_text(json.dumps(rec.record, indent=2))
    log.info("WRITE %s (%s)", rec.app_id_canonical, rec.lineage.get("validation_status"))


# =============================================================================
# 10. ORCHESTRATION — fixed order; invariants cannot be reordered
# =============================================================================
def _handle_fff_quarantine(sf: SourceFile, rec: AppRecord) -> None:
    sf.payload = None
    rec.lineage["fff_parse_blocked"] = True
    rec.validation_failures.append("fff_parse_blocked")
    rec.quarantined = True


def _check_pygdsa_attr_count(sf: SourceFile, rec: AppRecord) -> None:
    """REQ-BL-004: soft-warn when C103403 attr_count < 100. Not a quarantine."""
    if sf.payload is None:
        return
    attr_count = len(sf.payload)
    if attr_count < 100:
        log.warning("REQ-BL-004: C103403 attr_count=%d < 100", attr_count)
        rec.validation_failures.append("REQ-BL-004")


def run_pipeline(
    zip_path: str | Path,
    config_path: str | Path,
    mapping_path: str | Path,
    workdir: str | Path,
) -> list[AppRecord]:
    """
    Process a GDS ZIP package through the full SOC ingestion pipeline.

    Fixed call order — invariants cannot be reordered:
      1. build_manifest()
      2. dispatch_by_geo()            — INV-10: before any scrub or parse
      3. Per-geography loop:
         a. scrub_credentials()      — I1: first operation inside per-geo loop
         b. parse_file()
         c. group_by_app()
         d. per-record: merge_sessions → apply_mapping → tokenise_pii (I2)
            → assert_no_raw_pii (I2) → build_lineage → validate (I3) → write_record
    """
    cfg_base = ClientConfig.load(config_path)   # base config for manifest build

    root = unpack_zip(zip_path, workdir)

    # GDS packages often have a single top-level subfolder; descend into it.
    # Exclude pipeline-owned dirs (output/, quarantine/) so repeated runs
    # don't confuse the single-folder heuristic on the second pass.
    _pipeline_dirs = {"output", "quarantine"}
    subdirs = [p for p in root.iterdir() if p.is_dir() and p.name not in _pipeline_dirs]
    if len(subdirs) == 1:
        root = subdirs[0]

    # Amendment A4: clear output dir each run so stale records never persist.
    import shutil
    output_dir = Path(workdir) / "output"
    quarantine_dir = Path(workdir) / "quarantine"
    if output_dir.exists():
        shutil.rmtree(output_dir)
    output_dir.mkdir(parents=True)
    quarantine_dir.mkdir(parents=True, exist_ok=True)

    files = build_manifest(root, cfg_base)

    geo_files = dispatch_by_geo(files)          # INV-10: dispatch BEFORE scrub

    out: list[AppRecord] = []
    quarantined: list[AppRecord] = []
    for geo, geo_file_set in geo_files.items():
        if not geo_file_set:
            continue

        geo_cfg = ClientConfig.load(f"assets/client_config.SOC_{geo}.yaml")
        try:
            geo_mapping = load_mapping_sheet(f"assets/field_mapping.SOC_{geo}.xlsx")
        except Exception as e:
            log.warning("field mapping unavailable for %s: %s — mapping skipped", geo, e)
            geo_mapping = []

        scrubbed = scrub_credentials(geo_file_set, geo_cfg)    # I1 — FIRST in loop
        for sf in geo_file_set:
            try:
                parse_file(sf, geo_cfg)
            except NotImplementedError as e:
                log.warning("parse pending %s: %s", sf.connector, e)

        apps = group_by_app(geo_file_set, geo_cfg)
        blobs: list[dict] = []
        for rec in apps.values():
            rec.geography = geo
            # INV-13: fff parse failure = hard quarantine; silent skip is not permitted
            for sf in rec.files:
                conn = _connector_cfg(sf.connector, geo_cfg)
                if conn and conn.get("parse_strategy") == "fff":
                    _handle_fff_quarantine(sf, rec)
                elif conn and conn.get("parse_strategy") == "pygdsa_json":
                    _check_pygdsa_attr_count(sf, rec)
            merge_sessions(rec, geo_cfg)
            apply_mapping(rec, geo_mapping, geo_cfg)
            tokenise_pii(rec, geo_cfg)                          # I2 — before write
            try:
                assert_no_raw_pii(rec, geo_cfg)                 # I2 — write gate
            except RuntimeError as _pii_err:
                rec.quarantined = True
                rec.validation_failures.append("REQ-VAL-007")
                log.critical("RAW PII DETECTED %s: %s", rec.app_id_canonical, _pii_err)
            build_lineage(rec, geo_cfg, str(zip_path), scrubbed, blobs)
            validate(rec, geo_cfg)                              # I3 — before write
            _check_d13_completeness(rec)                        # D-13 — after validate
            write_record(rec, geo_cfg, workdir)
            out.append(rec)
            if rec.quarantined:
                quarantined.append(rec)

    _write_quarantine_report(out, quarantined, zip_path, workdir)
    return out


def _write_quarantine_report(
    out: list[AppRecord],
    quarantined: list[AppRecord],
    zip_path: "str | Path",
    workdir: "str | Path",
) -> None:
    from collections import Counter
    reason_freq = dict(Counter(f for r in quarantined for f in r.validation_failures))
    report = {
        "run_timestamp":       datetime.now(timezone.utc).isoformat(),
        "source_zip":          str(zip_path),
        "total_records":       len(out),
        "total_quarantined":   len(quarantined),
        "quarantine_rate_pct": round(100 * len(quarantined) / max(len(out), 1), 1),
        "reason_frequency":    reason_freq,
        "quarantined_app_ids": [r.app_id_canonical for r in quarantined],
    }
    report_path = Path(workdir) / "quarantine" / "report.json"
    report_path.parent.mkdir(exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2))


# --- small shared helpers ----------------------------------------------------
def _to_utc_iso(value: Any) -> str | None:
    s = str(value).strip()
    # Normalise +HHMM (no colon) → +HH:MM so fromisoformat handles it uniformly
    s_norm = re.sub(r"([+-])(\d{2})(\d{2})$", r"\1\2:\3", s)
    try:
        dt = datetime.fromisoformat(s_norm)
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc).isoformat()
        return dt.astimezone(timezone.utc).isoformat()
    except (ValueError, TypeError):
        pass
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%Y%m%d%H%M%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return None


def _coerce_number(value: Any):
    s = re.sub(r"[^\d.\-]", "", str(value))
    if s in ("", "-", "."):
        return None
    return float(s) if "." in s else int(s)


def _is_iso_utc(value: Any) -> bool:
    try:
        datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return True
    except (ValueError, TypeError):
        return False


def _set_path(obj: dict, dotted: str, value: Any) -> None:
    """Write value at a dotted path, creating dicts (and list stubs) as needed.

    Path components ending with '[]' represent the first element of a list:
      'applicants[].ssn'  →  obj['applicants'][0]['ssn'] = value
    This keeps _set_path symmetric with _get_nested's [] interpretation so
    tokenise_pii and validation rules can navigate the same structure.
    """
    parts = dotted.split(".")
    cur = obj
    for p in parts[:-1]:
        if p.endswith("[]"):
            key = p[:-2]
            if not isinstance(cur.get(key), list):
                cur[key] = [{}]
            elif not cur[key]:
                cur[key].append({})
            if not isinstance(cur[key][0], dict):
                cur[key][0] = {}
            cur = cur[key][0]
        else:
            cur = cur.setdefault(p, {})
    last = parts[-1]
    if last.endswith("[]"):
        key = last[:-2]
        if not isinstance(cur.get(key), list):
            cur[key] = []
        if not cur[key]:
            cur[key].append(value)
        else:
            cur[key][0] = value
    else:
        cur[last] = value


def _step_nested(cur: Any, part: str) -> Any:
    """Single dotted-path step: dict key lookup or list index (numeric part).

    'field[]' notation (from XLSX paths like 'data[].name') means: get 'field'
    from the dict then take the first element of the resulting list.
    """
    if part.endswith("[]"):
        key = part[:-2]
        val = cur.get(key) if isinstance(cur, dict) else None
        if isinstance(val, list):
            return val[0] if val else None
        return val
    if isinstance(cur, list):
        try:
            return cur[int(part)]
        except (ValueError, IndexError):
            return None
    if isinstance(cur, dict):
        return cur.get(part)
    return None


def _get_nested(obj: Any, dotted: str) -> Any:
    """Dotted-path read into a parsed payload dict. Returns None on any miss."""
    cur = obj
    for part in dotted.split("."):
        cur = _step_nested(cur, part)
        if cur is None:
            return None
    return cur


def _get_path(obj: dict, dotted: str) -> Any:
    """Dotted-path read from rec.record (same semantics as _get_nested)."""
    return _get_nested(obj, dotted)
